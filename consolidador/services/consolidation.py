from __future__ import annotations

import shutil
import tempfile
import re
import unicodedata
from datetime import date, datetime
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import SubmissionFile
from ..plan_types import plan_label
from .secure_storage import create_processing_copy, persist_file_for_storage


HEADER_FILL = PatternFill(fill_type="solid", fgColor="17324D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GUIDE_ROW_PHRASES = (
    "layout para importacao de dados",
    "uma inclusao por linha",
    "dados separados por",
    "campos obrigatorios",
    "tamanho max",
    "detalhe do campo",
    "obrigatorio",
    "somente para titulares",
    "opcoes:",
)


@dataclass(slots=True)
class ConsolidationBuildResult:
    plan_type: str
    consolidated_df: pd.DataFrame
    summary_rows: list[dict[str, object]]
    issues: list[dict[str, object]]
    consolidated_rows: int
    valid_files: int
    invalid_files: int
    base_reference: str
    base_headers_exact: list[str]
    had_header_adjustments: bool


def create_readable_copy(path: Path) -> Path:
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=path.suffix)
    temp_path = Path(temp_file.name)
    temp_file.close()
    shutil.copy2(path, temp_path)
    return temp_path


def read_table(path: Path, sheet_name: str | None) -> pd.DataFrame:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        last_error: Exception | None = None
        for encoding in ["utf-8", "utf-8-sig", "latin1", "cp1252"]:
            try:
                return pd.read_csv(path, dtype=object, sep=None, engine="python", encoding=encoding)
            except Exception as exc:
                last_error = exc
        raise ValueError(f"Nao foi possivel ler o CSV: {last_error}")

    target_sheet = sheet_name if sheet_name else 0
    try:
        return pd.read_excel(path, sheet_name=target_sheet, dtype=object)
    except ValueError:
        return pd.read_excel(path, sheet_name=0, dtype=object)


def read_excel_candidates(path: Path) -> list[tuple[str, pd.DataFrame]]:
    candidates: list[tuple[str, pd.DataFrame]] = []
    workbook = pd.ExcelFile(path)
    for candidate_sheet_name in workbook.sheet_names:
        try:
            dataframe = pd.read_excel(workbook, sheet_name=candidate_sheet_name, dtype=object)
        except Exception:
            continue
        candidates.append((str(candidate_sheet_name), dataframe))
    return candidates


def normalize_column_name(column_name: object) -> str:
    text = str(column_name).replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_headers(columns: Iterable[object]) -> list[str]:
    normalized_headers = [normalize_column_name(column) for column in columns]
    if len(set(normalized_headers)) != len(normalized_headers):
        raise ValueError("Colunas duplicadas encontradas apos a normalizacao dos nomes.")
    return normalized_headers


def extract_exact_headers(columns: Iterable[object]) -> list[str]:
    return [str(column).strip() for column in columns]


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    normalized = df.copy()
    normalized.columns = normalize_headers(normalized.columns)
    normalized = normalized.map(
        lambda value: pd.NA if isinstance(value, str) and not value.strip() else value
    )
    keep_columns: list[str] = []
    for column_name in normalized.columns:
        series = normalized[column_name]
        header_is_blank = column_name == "" or column_name.lower() in {"nan", "none", "unnamed"}
        if not header_is_blank or series.notna().any():
            keep_columns.append(column_name)
    normalized = normalized.loc[:, keep_columns]
    return normalized


def maybe_drop_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    return df.dropna(axis=0, how="all")


def excel_safe_value(value: object) -> object:
    if pd.isna(value):
        return None
    return value


def _clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def _strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(character for character in normalized if not unicodedata.combining(character))


def _clean_upper_text(value: object) -> str:
    return _strip_accents(_clean_text(value)).upper()


def canonical_cell_value(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value).strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def fold_text(value: object) -> str:
    text = canonical_cell_value(value)
    folded = unicodedata.normalize("NFKD", text)
    folded = "".join(character for character in folded if not unicodedata.combining(character))
    return folded.lower()


def is_date_like_header(header: str) -> bool:
    folded = fold_text(header)
    return any(
        token in folded
        for token in (
            "data",
            "nascimento",
            "admiss",
            "emissao",
            "vigencia",
            "vencimento",
            "inclusao",
            "exclusao",
        )
    )


def is_sex_like_header(header: str) -> bool:
    folded = fold_text(header)
    return "sexo" in folded


def is_cpf_like_header(header: str) -> bool:
    folded = fold_text(header)
    return folded == "cpf" or "cpf" in folded


def is_cep_like_header(header: str) -> bool:
    folded = fold_text(header)
    return "cep" in folded


def is_phone_like_header(header: str) -> bool:
    folded = fold_text(header)
    return any(token in folded for token in ("telefone", "celular", "fone", "whatsapp", "contato"))


def is_cns_like_header(header: str) -> bool:
    folded = fold_text(header)
    return folded == "cns" or "cartao nacional de saude" in folded


def is_rg_like_header(header: str) -> bool:
    folded = fold_text(header)
    return folded == "rg" or "identidade" in folded or "emissao rg" in folded


def format_date_output(value: object) -> object:
    if pd.isna(value):
        return pd.NA

    if isinstance(value, pd.Timestamp):
        return value.strftime("%d/%m/%Y")

    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")

    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    text = _clean_text(value)
    if not text:
        return pd.NA

    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return _clean_upper_text(text)
    return parsed.strftime("%d/%m/%Y")


def format_sex_output(value: object) -> object:
    if pd.isna(value):
        return pd.NA

    folded = fold_text(value)
    if not folded:
        return pd.NA
    if folded in {"m", "masc", "masculino", "male"}:
        return "M"
    if folded in {"f", "fem", "feminino", "female"}:
        return "F"
    return _clean_upper_text(value)


def format_cpf_output(value: object) -> object:
    if pd.isna(value):
        return pd.NA

    digits = re.sub(r"\D+", "", _clean_text(value))
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return _clean_upper_text(value)


def format_cep_output(value: object) -> object:
    if pd.isna(value):
        return pd.NA

    digits = re.sub(r"\D+", "", _clean_text(value))
    if len(digits) == 8:
        return f"{digits[:5]}-{digits[5:]}"
    return _clean_upper_text(value)


def format_phone_output(value: object) -> object:
    if pd.isna(value):
        return pd.NA

    digits = re.sub(r"\D+", "", _clean_text(value))
    if len(digits) == 11:
        return f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
    return _clean_upper_text(value)


def format_cns_output(value: object) -> object:
    if pd.isna(value):
        return pd.NA

    digits = re.sub(r"\D+", "", _clean_text(value))
    if len(digits) == 15:
        return digits
    return _clean_upper_text(value)


def format_rg_output(value: object) -> object:
    if pd.isna(value):
        return pd.NA

    cleaned = _clean_upper_text(value)
    if not cleaned:
        return pd.NA

    compact = re.sub(r"\s+", "", cleaned)
    if re.fullmatch(r"[0-9X.\-]+", compact):
        return compact
    return cleaned


def format_generic_output(value: object) -> object:
    if pd.isna(value):
        return pd.NA
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return format_date_output(value)
    if isinstance(value, str):
        cleaned = _clean_upper_text(value)
        return cleaned if cleaned else pd.NA
    return value


def standardize_output_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    standardized = dataframe.copy()
    for column_name in standardized.columns:
        if is_cpf_like_header(column_name):
            standardized[column_name] = standardized[column_name].map(format_cpf_output)
            continue
        if is_cep_like_header(column_name):
            standardized[column_name] = standardized[column_name].map(format_cep_output)
            continue
        if is_phone_like_header(column_name):
            standardized[column_name] = standardized[column_name].map(format_phone_output)
            continue
        if is_cns_like_header(column_name):
            standardized[column_name] = standardized[column_name].map(format_cns_output)
            continue
        if is_rg_like_header(column_name):
            standardized[column_name] = standardized[column_name].map(format_rg_output)
            continue
        if is_sex_like_header(column_name):
            standardized[column_name] = standardized[column_name].map(format_sex_output)
            continue
        if is_date_like_header(column_name):
            standardized[column_name] = standardized[column_name].map(format_date_output)
            continue
        standardized[column_name] = standardized[column_name].map(format_generic_output)
    return standardized


def build_row_signatures(dataframe: pd.DataFrame) -> set[tuple[str, ...]]:
    signatures: set[tuple[str, ...]] = set()
    for row in dataframe.itertuples(index=False, name=None):
        signature = tuple(canonical_cell_value(value) for value in row)
        if any(signature):
            signatures.add(signature)
    return signatures


def build_compact_row_signatures(dataframe: pd.DataFrame) -> set[tuple[str, ...]]:
    signatures: set[tuple[str, ...]] = set()
    for row in dataframe.itertuples(index=False, name=None):
        signature = tuple(canonical_cell_value(value) for value in row if canonical_cell_value(value))
        if signature:
            signatures.add(signature)
    return signatures


def drop_reference_rows(
    dataframe: pd.DataFrame,
    reference_signatures: set[tuple[str, ...]],
) -> tuple[pd.DataFrame, int]:
    if dataframe.empty or not reference_signatures:
        return dataframe, 0

    keep_indexes: list[int] = []
    removed_rows = 0
    for row_index, row in zip(dataframe.index, dataframe.itertuples(index=False, name=None)):
        signature = tuple(canonical_cell_value(value) for value in row)
        if signature in reference_signatures:
            removed_rows += 1
            continue
        keep_indexes.append(row_index)

    filtered = dataframe.loc[keep_indexes].reset_index(drop=True)
    return filtered, removed_rows


def row_looks_like_guide(
    signature: tuple[str, ...],
    *,
    base_headers_exact: list[str],
) -> bool:
    non_empty_values = [value for value in signature if value]
    if not non_empty_values:
        return False

    folded_values = [fold_text(value) for value in non_empty_values]
    joined_values = " | ".join(folded_values)

    if any(phrase in joined_values for phrase in GUIDE_ROW_PHRASES):
        return True

    first_value = folded_values[0]
    folded_headers = {fold_text(header) for header in base_headers_exact}
    if first_value in folded_headers and len(non_empty_values) <= 4:
        return True

    return False


def drop_guide_rows(
    dataframe: pd.DataFrame,
    *,
    reference_signatures: set[tuple[str, ...]],
    compact_reference_signatures: set[tuple[str, ...]],
    base_headers_exact: list[str],
) -> tuple[pd.DataFrame, int]:
    if dataframe.empty:
        return dataframe, 0

    keep_indexes: list[int] = []
    removed_rows = 0

    for row_index, row in zip(dataframe.index, dataframe.itertuples(index=False, name=None)):
        signature = tuple(canonical_cell_value(value) for value in row)
        compact_signature = tuple(value for value in signature if value)
        if (
            signature in reference_signatures
            or compact_signature in compact_reference_signatures
            or row_looks_like_guide(signature, base_headers_exact=base_headers_exact)
        ):
            removed_rows += 1
            continue
        keep_indexes.append(row_index)

    filtered = dataframe.loc[keep_indexes].reset_index(drop=True)
    return filtered, removed_rows


def estimate_candidate_data_rows(
    dataframe: pd.DataFrame,
    *,
    base_columns: list[str] | None,
    base_headers_exact: list[str] | None,
    base_reference_rows: set[tuple[str, ...]],
    base_reference_compact_rows: set[tuple[str, ...]],
    ignore_empty: bool,
) -> int:
    working = normalize_dataframe(dataframe)
    if ignore_empty:
        working = maybe_drop_empty_rows(working)

    file_columns = list(working.columns)
    if not file_columns:
        return -1

    if base_columns is not None:
        if base_headers_exact:
            working, _, unexpected_columns, _ = align_dataframe_to_base(
                working,
                base_headers_exact=base_headers_exact,
                base_columns=base_columns,
            )
            if unexpected_columns:
                return -1
            working.columns = base_headers_exact
            working, _ = drop_reference_rows(working, base_reference_rows)
            working, _ = drop_guide_rows(
                working,
                reference_signatures=base_reference_rows,
                compact_reference_signatures=base_reference_compact_rows,
                base_headers_exact=base_headers_exact,
            )
        elif len(file_columns) != len(base_columns):
            return -1

    if ignore_empty:
        working = maybe_drop_empty_rows(working)

    return len(working)


def choose_best_sheet_name(
    path: Path,
    *,
    base_columns: list[str] | None,
    base_headers_exact: list[str] | None,
    base_reference_rows: set[tuple[str, ...]],
    base_reference_compact_rows: set[tuple[str, ...]],
    ignore_empty: bool,
) -> str | None:
    if path.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
        return None

    best_sheet_name: str | None = None
    best_score = -1

    for candidate_sheet_name, candidate_df in read_excel_candidates(path):
        try:
            score = estimate_candidate_data_rows(
                candidate_df,
                base_columns=base_columns,
                base_headers_exact=base_headers_exact,
                base_reference_rows=base_reference_rows,
                base_reference_compact_rows=base_reference_compact_rows,
                ignore_empty=ignore_empty,
            )
        except Exception:
            continue

        if score > best_score:
            best_score = score
            best_sheet_name = candidate_sheet_name

    return best_sheet_name


def align_dataframe_to_base(
    dataframe: pd.DataFrame,
    *,
    base_headers_exact: list[str],
    base_columns: list[str],
) -> tuple[pd.DataFrame, list[str], list[str], bool]:
    upload_columns = list(dataframe.columns)
    allow_positional_fallback = len(upload_columns) == len(base_columns)
    used_indexes: set[int] = set()
    aligned_columns: dict[str, pd.Series] = {}
    missing_columns: list[str] = []
    unexpected_columns: list[str] = []
    used_positional_fallback = False

    for base_index, expected_normalized in enumerate(base_columns):
        matched_index: int | None = None

        for upload_index, upload_column in enumerate(upload_columns):
            if upload_index in used_indexes:
                continue
            if upload_column == expected_normalized:
                matched_index = upload_index
                break

        if (
            matched_index is None
            and allow_positional_fallback
            and base_index < len(upload_columns)
            and base_index not in used_indexes
        ):
            matched_index = base_index
            used_positional_fallback = True

        target_header = base_headers_exact[base_index]
        if matched_index is None:
            aligned_columns[target_header] = pd.Series([pd.NA] * len(dataframe), index=dataframe.index, dtype=object)
            missing_columns.append(target_header)
        else:
            used_indexes.add(matched_index)
            aligned_columns[target_header] = dataframe.iloc[:, matched_index]

    for upload_index, upload_column in enumerate(upload_columns):
        if upload_index in used_indexes:
            continue
        series = dataframe.iloc[:, upload_index]
        if series.notna().any():
            unexpected_columns.append(upload_column)

    aligned_dataframe = pd.DataFrame(aligned_columns, index=dataframe.index)
    return aligned_dataframe, missing_columns, unexpected_columns, used_positional_fallback


def build_consolidated_workbook(
    files: Iterable[SubmissionFile],
    *,
    plan_type: str,
    sheet_name: str | None,
    base_template_path: Path | None,
    remove_duplicates: bool,
    ignore_empty: bool,
    include_source_columns: bool,
) -> ConsolidationBuildResult:
    base_columns: list[str] | None = None
    base_headers_exact: list[str] | None = None
    base_reference_rows: set[tuple[str, ...]] = set()
    base_reference_compact_rows: set[tuple[str, ...]] = set()
    base_reference = ""
    consolidated_parts: list[pd.DataFrame] = []
    issues: list[dict[str, object]] = []
    summary_rows: list[dict[str, object]] = []
    valid_files = 0
    invalid_files = 0
    had_header_adjustments = False

    if base_template_path and base_template_path.exists():
        template_copy = create_readable_copy(base_template_path)
        try:
            template_df = read_table(template_copy, sheet_name=sheet_name)
            base_headers_exact = extract_exact_headers(template_df.columns)
            base_columns = normalize_headers(base_headers_exact)
            normalized_template_df = normalize_dataframe(template_df)
            normalized_template_df.columns = base_headers_exact
            normalized_template_df = maybe_drop_empty_rows(normalized_template_df)
            base_reference_rows = build_row_signatures(normalized_template_df)
            base_reference_compact_rows = build_compact_row_signatures(normalized_template_df)
            base_reference = base_template_path.name
        finally:
            template_copy.unlink(missing_ok=True)

    for item in files:
        path = Path(item.stored_path)
        working_copy: Path | None = None
        try:
            if not path.exists():
                raise FileNotFoundError("Arquivo salvo nao encontrado no armazenamento.")
            source_suffix = Path(item.original_name or item.stored_name).suffix or ".tmp"
            working_copy = create_processing_copy(path, suffix=source_suffix)

            effective_sheet_name = sheet_name
            if not effective_sheet_name:
                detected_sheet_name = choose_best_sheet_name(
                    working_copy,
                    base_columns=base_columns,
                    base_headers_exact=base_headers_exact,
                    base_reference_rows=base_reference_rows,
                    base_reference_compact_rows=base_reference_compact_rows,
                    ignore_empty=ignore_empty,
                )
                effective_sheet_name = detected_sheet_name or None

            dataframe = read_table(working_copy, sheet_name=effective_sheet_name)
            original_headers = extract_exact_headers(dataframe.columns)
            normalized_headers = normalize_headers(original_headers)
            header_adjusted_for_file = False
            dataframe = normalize_dataframe(dataframe)
            original_rows = len(dataframe)

            if ignore_empty:
                dataframe = maybe_drop_empty_rows(dataframe)

            file_columns = list(dataframe.columns)

            if not file_columns:
                raise ValueError("Nenhuma coluna util foi encontrada apos a limpeza do arquivo.")

            if base_columns is None:
                base_columns = file_columns
                base_headers_exact = original_headers
                base_reference = item.original_name
            else:
                if base_headers_exact:
                    (
                        dataframe,
                        missing_columns,
                        unexpected_columns,
                        used_positional_fallback,
                    ) = align_dataframe_to_base(
                        dataframe,
                        base_headers_exact=base_headers_exact,
                        base_columns=base_columns,
                    )
                    if unexpected_columns:
                        invalid_files += 1
                        issues.append(
                            {
                                "arquivo": item.original_name,
                                "tipo": plan_label(plan_type),
                                "submissao_id": item.submission_id,
                                "status": "ignorado",
                                "motivo": "Colunas extras com dados nao puderam ser alinhadas com a planilha base.",
                                "colunas_arquivo": " | ".join(file_columns),
                                "colunas_esperadas": " | ".join(base_columns),
                            }
                        )
                        summary_rows.append(
                            {
                                "arquivo": item.original_name,
                                "tipo": plan_label(plan_type),
                                "enviado_por": item.sender_name or item.sender_email or "-",
                                "submissao": item.submission_id,
                                "linhas_lidas": original_rows,
                                "linhas_utilizadas": 0,
                                "status": "Ignorado",
                                "motivo": "Colunas extras nao alinhadas com a base",
                            }
                        )
                        continue

                    if missing_columns or used_positional_fallback:
                        had_header_adjustments = True
                        header_adjusted_for_file = True

                    file_columns = list(dataframe.columns)
                elif len(file_columns) != len(base_columns):
                    invalid_files += 1
                    issues.append(
                        {
                            "arquivo": item.original_name,
                            "tipo": plan_label(plan_type),
                            "submissao_id": item.submission_id,
                            "status": "ignorado",
                            "motivo": "Quantidade de colunas diferente da planilha base.",
                            "colunas_arquivo": " | ".join(file_columns),
                            "colunas_esperadas": " | ".join(base_columns),
                        }
                    )
                    summary_rows.append(
                        {
                            "arquivo": item.original_name,
                            "tipo": plan_label(plan_type),
                            "enviado_por": item.sender_name or item.sender_email or "-",
                            "submissao": item.submission_id,
                            "linhas_lidas": original_rows,
                            "linhas_utilizadas": 0,
                            "status": "Ignorado",
                            "motivo": "Quantidade de colunas diferente da base",
                        }
                    )
                    continue

                if not base_headers_exact and normalized_headers != base_columns:
                    had_header_adjustments = True
                    header_adjusted_for_file = True

            if base_headers_exact:
                dataframe.columns = base_headers_exact
                file_columns = base_headers_exact
                dataframe, removed_reference_rows = drop_reference_rows(dataframe, base_reference_rows)
                dataframe, removed_guide_rows = drop_guide_rows(
                    dataframe,
                    reference_signatures=base_reference_rows,
                    compact_reference_signatures=base_reference_compact_rows,
                    base_headers_exact=base_headers_exact,
                )
                removed_reference_rows += removed_guide_rows
            else:
                removed_reference_rows = 0

            if ignore_empty:
                dataframe = maybe_drop_empty_rows(dataframe)

            rows_after_cleanup = len(dataframe)

            if removed_reference_rows:
                had_header_adjustments = True

            if dataframe.empty:
                invalid_files += 1
                issues.append(
                    {
                        "arquivo": item.original_name,
                        "tipo": plan_label(plan_type),
                        "submissao_id": item.submission_id,
                        "status": "ignorado",
                        "motivo": "Nenhuma linha de dados real foi encontrada apos remover as linhas-guia da planilha base.",
                        "colunas_arquivo": " | ".join(file_columns),
                        "colunas_esperadas": " | ".join(base_columns or []),
                    }
                )
                summary_rows.append(
                    {
                        "arquivo": item.original_name,
                        "tipo": plan_label(plan_type),
                        "enviado_por": item.sender_name or item.sender_email or "-",
                        "submissao": item.submission_id,
                        "linhas_lidas": original_rows,
                        "linhas_utilizadas": 0,
                        "status": "Ignorado",
                        "motivo": "Somente linhas-guia/modelo foram encontradas",
                    }
                )
                continue

            if include_source_columns and not base_template_path:
                dataframe.insert(0, "Arquivo_Origem", item.original_name)
                dataframe.insert(1, "Submissao_ID", item.submission_id)
                dataframe.insert(2, "Enviado_Em", item.submitted_at)
                dataframe.insert(3, "Enviado_Por", item.sender_name or item.sender_email or "")

            consolidated_parts.append(dataframe)
            valid_files += 1
            summary_rows.append(
                {
                    "arquivo": item.original_name,
                    "tipo": plan_label(plan_type),
                    "enviado_por": item.sender_name or item.sender_email or "-",
                    "submissao": item.submission_id,
                    "linhas_lidas": original_rows,
                    "linhas_utilizadas": rows_after_cleanup,
                    "status": "OK",
                    "motivo": (
                        "Cabecalhos ajustados e linhas-guia removidas"
                        if header_adjusted_for_file and removed_reference_rows
                        else (
                            "Cabecalhos ajustados para a base"
                            if header_adjusted_for_file and base_template_path
                            else (
                                "Linhas-guia da base removidas"
                                if removed_reference_rows
                                else ("Base definida" if item.original_name == base_reference else "")
                            )
                        )
                    ),
                }
            )
        except Exception as exc:
            invalid_files += 1
            issues.append(
                {
                    "arquivo": item.original_name,
                    "tipo": plan_label(plan_type),
                    "submissao_id": item.submission_id,
                    "status": "erro",
                    "motivo": str(exc),
                    "colunas_arquivo": "",
                    "colunas_esperadas": " | ".join(base_columns or []),
                }
            )
            summary_rows.append(
                {
                    "arquivo": item.original_name,
                    "tipo": plan_label(plan_type),
                    "enviado_por": item.sender_name or item.sender_email or "-",
                    "submissao": item.submission_id,
                    "linhas_lidas": 0,
                    "linhas_utilizadas": 0,
                    "status": "Erro de leitura",
                    "motivo": str(exc),
                }
            )
        finally:
            if working_copy is not None:
                working_copy.unlink(missing_ok=True)

    if consolidated_parts:
        consolidated_df = pd.concat(consolidated_parts, ignore_index=True)
        consolidated_df = standardize_output_dataframe(consolidated_df)
        if remove_duplicates:
            consolidated_df = consolidated_df.drop_duplicates(ignore_index=True)
    else:
        consolidated_df = pd.DataFrame()

    summary_rows.insert(
        0,
        {
            "arquivo": "TOTAL",
            "tipo": plan_label(plan_type),
            "enviado_por": "",
            "submissao": "",
            "linhas_lidas": int(sum(int(row["linhas_lidas"]) for row in summary_rows)),
            "linhas_utilizadas": int(sum(int(row["linhas_utilizadas"]) for row in summary_rows)),
            "status": f"Validos: {valid_files} | Inconsistentes: {invalid_files}",
            "motivo": base_reference or "Nenhuma planilha base definida",
        },
    )

    if not issues:
        issues.append(
            {
                "arquivo": "",
                "tipo": plan_label(plan_type),
                "submissao_id": "",
                "status": "ok",
                "motivo": "Nenhuma inconsistencia encontrada.",
                "colunas_arquivo": "",
                "colunas_esperadas": "",
            }
        )

    return ConsolidationBuildResult(
        plan_type=plan_type,
        consolidated_df=consolidated_df,
        summary_rows=summary_rows,
        issues=issues,
        consolidated_rows=len(consolidated_df),
        valid_files=valid_files,
        invalid_files=invalid_files,
        base_reference=base_reference,
        base_headers_exact=base_headers_exact or list(consolidated_df.columns),
        had_header_adjustments=had_header_adjustments,
    )


def clear_existing_values(worksheet, max_column: int) -> None:
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=max_column):
        for cell in row:
            cell.value = None


def ensure_row_style(worksheet, row_number: int, max_column: int) -> None:
    if row_number <= worksheet.max_row:
        return
    source_row = row_number - 1 if row_number > 2 else 2
    if source_row > worksheet.max_row:
        source_row = worksheet.max_row if worksheet.max_row >= 2 else 0
    if source_row < 1:
        return

    for column_index in range(1, max_column + 1):
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=row_number, column=column_index)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[row_number]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden


def write_consolidation_output(
    result: ConsolidationBuildResult,
    output_path: Path,
    *,
    main_sheet_name: str = "Consolidado",
    base_template_path: Path | None = None,
    sheet_name: str | None = None,
) -> None:
    if base_template_path and base_template_path.exists():
        template_copy = create_readable_copy(base_template_path)
        workbook = load_workbook(template_copy)
        template_copy.unlink(missing_ok=True)
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        for extra_sheet_name in list(workbook.sheetnames):
            if extra_sheet_name != worksheet.title:
                del workbook[extra_sheet_name]
    else:
        workbook = Workbook()
        worksheet = workbook.active

    worksheet.title = main_sheet_name[:31]
    if result.base_headers_exact:
        for column_index, header in enumerate(result.base_headers_exact, start=1):
            worksheet.cell(row=1, column=column_index, value=header)
    clear_existing_values(worksheet, max_column=max(len(result.base_headers_exact), worksheet.max_column))

    for row_index, row in enumerate(result.consolidated_df.itertuples(index=False, name=None), start=2):
        ensure_row_style(worksheet, row_index, max(len(result.base_headers_exact), len(row)))
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=excel_safe_value(value))

    if "Resumo" in workbook.sheetnames:
        del workbook["Resumo"]
    if "Inconsistencias" in workbook.sheetnames:
        del workbook["Inconsistencias"]

    summary_sheet = workbook.create_sheet("Resumo")
    issues_sheet = workbook.create_sheet("Inconsistencias")

    for target_sheet, rows in ((summary_sheet, result.summary_rows), (issues_sheet, result.issues)):
        dataframe = pd.DataFrame(rows)
        headers = list(dataframe.columns)
        for column_index, value in enumerate(headers, start=1):
            target_sheet.cell(row=1, column=column_index, value=value)
        for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
            for column_index, value in enumerate(row, start=1):
                target_sheet.cell(row=row_index, column=column_index, value=excel_safe_value(value))

    for worksheet_item in workbook.worksheets:
        worksheet_item.freeze_panes = "A2"
        if worksheet_item.max_row and worksheet_item.max_column:
            worksheet_item.auto_filter.ref = worksheet_item.dimensions
        for cell in worksheet_item[1]:
            if worksheet_item.title in {"Resumo", "Inconsistencias"} or not base_template_path:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
        if base_template_path and worksheet_item.title == main_sheet_name[:31]:
            continue
        for column_cells in worksheet_item.columns:
            length = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                length = max(length, len(str(cell.value)))
            worksheet_item.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                max(length + 2, 12),
                48,
            )

    workbook.save(output_path)


def write_consolidation_output_for_storage(
    result: ConsolidationBuildResult,
    target_path: Path,
    *,
    main_sheet_name: str = "Consolidado",
    base_template_path: Path | None = None,
    sheet_name: str | None = None,
) -> Path:
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=target_path.suffix or ".xlsx")
    temp_path = Path(temp_file.name)
    temp_file.close()
    write_consolidation_output(
        result,
        temp_path,
        main_sheet_name=main_sheet_name,
        base_template_path=base_template_path,
        sheet_name=sheet_name,
    )
    return persist_file_for_storage(temp_path, target_path)
